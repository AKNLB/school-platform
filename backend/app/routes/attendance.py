from flask import Blueprint, request, jsonify
import csv
import io
from datetime import datetime

from openpyxl import load_workbook

from app.audit import log_audit
import app.bridge as bridge
from app.decorators import (
    ROLE_ADMIN,
    ROLE_TEACHER,
    current_school_id,
    roles_required,
    school_context_required,
)

attendance_bp = Blueprint("attendance_bp", __name__)


def _safe_date(value):
    if value in (None, ""):
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _normalize_header(value):
    return str(value or "").strip().lower()


def _clean_cell(row, key, default=""):
    value = row.get(key, default)
    if value is None:
        return default
    return str(value).strip()


def _normalize_status(value):
    raw = str(value or "present").strip().lower()

    mapping = {
        "present": "present",
        "p": "present",
        "absent": "absent",
        "a": "absent",
        "late": "late",
        "l": "late",
        "excused": "excused",
        "e": "excused",
        "excuse": "excused",
    }

    return mapping.get(raw, "")


def _parse_import_date(value):
    raw = str(value or "").strip()
    if not raw:
        return None

    try:
        return bridge.parse_date(raw)
    except Exception:
        pass

    formats = [
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y/%m/%d",
        "%m-%d-%Y",
        "%d-%m-%Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt).date()
        except Exception:
            continue

    raise ValueError("Invalid date")


def _rows_from_csv(file):
    raw = file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))

    if not reader.fieldnames:
        raise ValueError("CSV has no header row")

    rows = []
    for row in reader:
        rows.append({str(k).strip().lower(): v for k, v in row.items() if k is not None})

    return [_normalize_header(h) for h in reader.fieldnames if h], rows


def _rows_from_xlsx(file):
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        raise ValueError("Excel file is empty")

    headers = [_normalize_header(h) for h in rows_iter[0]]
    if not any(headers):
        raise ValueError("Excel file has no header row")

    rows = []
    for values in rows_iter[1:]:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[idx] if idx < len(values) else ""
        rows.append(row)

    return headers, rows


def _load_import_rows(file):
    filename = (file.filename or "").lower()

    if filename.endswith(".csv"):
        headers, rows = _rows_from_csv(file)
        return "csv", headers, rows

    if filename.endswith(".xlsx"):
        headers, rows = _rows_from_xlsx(file)
        return "xlsx", headers, rows

    raise ValueError("Only .csv and .xlsx files are supported")


@attendance_bp.route("/api/s/<slug>/attendance", methods=["GET", "POST", "PUT"])
@attendance_bp.route("/api/attendance", methods=["GET", "POST", "PUT"])
@school_context_required
@roles_required(ROLE_ADMIN, ROLE_TEACHER)
def attendance(slug=None):
    Attendance = bridge.Attendance
    Student = bridge.Student
    db = bridge.db

    if request.method == "GET":
        grade = request.args.get("grade", type=int)
        day = request.args.get("date")

        q = Attendance.query.filter_by(school_id=current_school_id())

        if grade is not None:
            q = q.join(Student, Student.id == Attendance.student_id).filter(
                Student.school_id == current_school_id(),
                Student.grade == grade,
            )

        if day:
            q = q.filter(Attendance.date == bridge.parse_date(day))

        recs = q.order_by(Attendance.date.desc(), Attendance.id.desc()).limit(300).all()

        return jsonify(
            [
                {
                    "id": r.id,
                    "date": _safe_date(r.date),
                    "student_id": r.student_id,
                    "student_name": r.student.name if r.student else None,
                    "grade": r.student.grade if r.student else None,
                    "status": r.status,
                    "note": r.note or "",
                }
                for r in recs
            ]
        ), 200

    payload = request.get_json(silent=True) or {}

    if request.method == "POST":
        if not payload.get("student_id") or not payload.get("date"):
            return jsonify({"error": "student_id and date required"}), 400

        student = Student.query.filter_by(
            id=int(payload["student_id"]),
            school_id=current_school_id(),
        ).first()

        if not student:
            return jsonify({"error": "Student not found"}), 404

        status = _normalize_status(payload.get("status") or "present")
        if not status:
            return jsonify({"error": "Invalid status. Use present, absent, late, or excused."}), 400

        r = Attendance(
            school_id=current_school_id(),
            student_id=student.id,
            date=bridge.parse_date(payload["date"]),
            status=status,
            note=payload.get("note"),
        )
        db.session.add(r)
        db.session.commit()

        log_audit(
            module="attendance",
            action="create",
            entity_type="attendance",
            entity_id=r.id,
            entity_label=f"{student.name} - {_safe_date(r.date)}",
            details={
                "student_id": student.id,
                "student_name": student.name,
                "date": _safe_date(r.date),
                "status": r.status,
                "note": r.note,
            },
        )

        return jsonify({"id": r.id}), 201

    items = payload if isinstance(payload, list) else [payload]

    updated = 0
    created = 0
    skipped = 0

    for data in items:
        if not isinstance(data, dict):
            skipped += 1
            continue

        sid = data.get("student_id")
        day = data.get("date")
        if not sid or not day:
            skipped += 1
            continue

        sid = int(sid)
        dt = bridge.parse_date(day)

        student = Student.query.filter_by(
            id=sid,
            school_id=current_school_id(),
        ).first()

        if not student:
            skipped += 1
            continue

        r = Attendance.query.filter_by(
            school_id=current_school_id(),
            student_id=sid,
            date=dt,
        ).first()

        if r is None:
            r = Attendance(
                school_id=current_school_id(),
                student_id=sid,
                date=dt,
            )
            db.session.add(r)
            created += 1
        else:
            updated += 1

        if "status" in data and data["status"] is not None:
            status = _normalize_status(data.get("status"))
            if status:
                r.status = status
        if "note" in data:
            r.note = data.get("note") or None

    db.session.commit()

    log_audit(
        module="attendance",
        action="bulk_save",
        entity_type="attendance",
        entity_id=None,
        entity_label="Bulk Attendance Save",
        details={
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "items_count": len(items),
        },
    )

    return jsonify({
        "message": "Attendance saved",
        "created": created,
        "updated": updated,
        "skipped": skipped,
    }), 200


@attendance_bp.route("/api/s/<slug>/attendance/import", methods=["POST"])
@attendance_bp.route("/api/attendance/import", methods=["POST"])
@school_context_required
@roles_required(ROLE_ADMIN, ROLE_TEACHER)
@bridge.limiter.limit("10 per hour")
def import_attendance(slug=None):
    Attendance = bridge.Attendance
    Student = bridge.Student
    db = bridge.db
    sid = current_school_id()

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "Import file is required"}), 400

    try:
        file_type, headers, import_rows = _load_import_rows(file)
    except UnicodeDecodeError:
        return jsonify({"error": "Could not read CSV. Please save it as UTF-8 CSV."}), 400
    except Exception as e:
        return jsonify({"error": str(e) or "Could not read import file"}), 400

    required = ["student_id", "date", "status"]
    normalized_headers = {h: h for h in headers if h}
    missing = [h for h in required if h not in normalized_headers]

    if missing:
        return jsonify({
            "error": f"Missing required column(s): {', '.join(missing)}",
            "required_columns": required,
            "optional_columns": ["note"],
            "allowed_statuses": ["present", "absent", "late", "excused"],
        }), 400

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for index, row in enumerate(import_rows, start=2):
        row = {str(k).strip().lower(): v for k, v in row.items() if k is not None}

        student_id_raw = _clean_cell(row, "student_id")
        date_raw = _clean_cell(row, "date")
        status_raw = _clean_cell(row, "status")
        note = _clean_cell(row, "note") or None

        if not student_id_raw or not date_raw:
            skipped += 1
            errors.append({
                "row": index,
                "error": "student_id and date are required",
            })
            continue

        try:
            student_id = int(float(student_id_raw))
        except Exception:
            skipped += 1
            errors.append({
                "row": index,
                "student_id": student_id_raw,
                "error": "student_id must be a number",
            })
            continue

        try:
            dt = _parse_import_date(date_raw)
        except Exception:
            skipped += 1
            errors.append({
                "row": index,
                "student_id": student_id,
                "error": "Invalid date. Use YYYY-MM-DD, MM/DD/YYYY, or DD/MM/YYYY.",
            })
            continue

        status = _normalize_status(status_raw)
        if not status:
            skipped += 1
            errors.append({
                "row": index,
                "student_id": student_id,
                "date": _safe_date(dt),
                "error": "Invalid status. Use present, absent, late, or excused.",
            })
            continue

        student = Student.query.filter_by(
            id=student_id,
            school_id=sid,
        ).first()

        if not student:
            skipped += 1
            errors.append({
                "row": index,
                "student_id": student_id,
                "error": "Student not found",
            })
            continue

        rec = Attendance.query.filter_by(
            school_id=sid,
            student_id=student_id,
            date=dt,
        ).first()

        if rec:
            rec.status = status
            rec.note = note
            updated += 1
        else:
            rec = Attendance(
                school_id=sid,
                student_id=student_id,
                date=dt,
                status=status,
                note=note,
            )
            db.session.add(rec)
            created += 1

    db.session.commit()

    log_audit(
        module="attendance",
        action="bulk_import",
        entity_type="attendance_import",
        entity_id=None,
        entity_label="Bulk Attendance Import",
        details={
            "file_type": file_type,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors_count": len(errors),
        },
    )

    return jsonify({
        "message": "Attendance import completed",
        "file_type": file_type,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:100],
    }), 200


@attendance_bp.route("/api/s/<slug>/attendance/student/<int:student_id>", methods=["GET"])
@attendance_bp.route("/api/attendance/student/<int:student_id>", methods=["GET"])
@school_context_required
@roles_required(ROLE_ADMIN, ROLE_TEACHER)
def attendance_student_history(student_id: int, slug=None):
    Attendance = bridge.Attendance
    Student = bridge.Student

    limit = int(request.args.get("limit", 30))

    student = Student.query.filter_by(
        id=student_id,
        school_id=current_school_id(),
    ).first()

    if not student:
        return jsonify({"error": "Student not found"}), 404

    rows = (
        Attendance.query
        .filter_by(
            school_id=current_school_id(),
            student_id=student_id,
        )
        .order_by(Attendance.date.desc(), Attendance.id.desc())
        .limit(limit)
        .all()
    )

    return jsonify([
        {
            "id": r.id,
            "student_id": r.student_id,
            "date": _safe_date(r.date),
            "status": r.status,
            "note": r.note or "",
        }
        for r in rows
    ]), 200


@attendance_bp.route("/api/s/<slug>/attendance/report", methods=["GET"])
@attendance_bp.route("/api/attendance/report", methods=["GET"])
@school_context_required
@roles_required(ROLE_ADMIN, ROLE_TEACHER)
def api_attendance_report(slug=None):
    Attendance = bridge.Attendance

    dt_from = request.args.get("from")
    dt_to = request.args.get("to")

    q = Attendance.query.filter_by(school_id=current_school_id())

    if dt_from:
        q = q.filter(Attendance.date >= bridge.parse_date(dt_from))
    if dt_to:
        q = q.filter(Attendance.date <= bridge.parse_date(dt_to))

    rows = q.order_by(Attendance.date.asc()).all()

    records = [
        {
            "date": _safe_date(a.date),
            "student_id": a.student_id,
            "status": a.status,
            "note": a.note,
        }
        for a in rows
    ]

    present = sum(1 for r in records if (r["status"] or "").lower() == "present")
    absent = sum(1 for r in records if (r["status"] or "").lower() == "absent")
    late = sum(1 for r in records if (r["status"] or "").lower() == "late")
    excused = sum(1 for r in records if (r["status"] or "").lower() == "excused")

    return jsonify({
        "records": records,
        "summary": {
            "present": present,
            "absent": absent,
            "late": late,
            "excused": excused,
            "total_marked": len(records),
        }
    }), 200