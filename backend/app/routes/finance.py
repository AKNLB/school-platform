from flask import Blueprint, request, jsonify, send_file
import io
import csv
from openpyxl import load_workbook

from app.audit import log_audit
import app.bridge as bridge
from app.decorators import (
    ROLE_ADMIN,
    current_school_id,
    roles_required,
    school_context_required,
)

finance_bp = Blueprint("finance_bp", __name__)

def _money(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0

def _clean_cell(row, key, default=""):
    value = row.get(key, default)
    if value is None:
        return default
    return str(value).strip()


def _to_float(value):
    try:
        raw = str(value or "0").replace("$", "").replace(",", "").strip()
        return float(raw or 0)
    except Exception:
        return 0.0


def _normalize_header(value):
    return str(value or "").strip().lower()


def _rows_from_csv(file):
    raw = file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))

    if not reader.fieldnames:
        raise ValueError("CSV has no header row")

    rows = []
    for row in reader:
        rows.append({str(k).strip().lower(): v for k, v in row.items() if k is not None})

    return [_normalize_header(h) for h in reader.fieldnames if h], rows


def _rows_from_xlsx(file):
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        raise ValueError("Excel file is empty")

    headers = [_normalize_header(h) for h in rows_iter[0]]
    if not any(headers):
        raise ValueError("Excel file has no header row")

    rows = []
    for values in rows_iter[1:]:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[idx] if idx < len(values) else ""
        rows.append(row)

    return headers, rows


def _load_import_rows(file):
    filename = (file.filename or "").lower()

    if filename.endswith(".csv"):
        headers, rows = _rows_from_csv(file)
        return "csv", headers, rows

    if filename.endswith(".xlsx"):
        headers, rows = _rows_from_xlsx(file)
        return "xlsx", headers, rows

    raise ValueError("Only .csv and .xlsx files are supported")

@finance_bp.route("/api/s/<slug>/tuition/<int:student_id>", methods=["GET"])
@finance_bp.route("/api/tuition/<int:student_id>", methods=["GET"])
@school_context_required
@roles_required(ROLE_ADMIN)
def get_tuition_info(student_id, slug=None):
    Student = bridge.Student
    TuitionInfo = bridge.TuitionInfo

    student = Student.query.filter_by(
        id=student_id,
        school_id=current_school_id()
    ).first()

    if not student:
        return jsonify({"error": "Student not found"}), 404

    term = request.args.get("term")

    q = TuitionInfo.query.filter_by(
        school_id=current_school_id(),
        student_id=student_id
    )

    if term:
        q = q.filter_by(term=term)

    tuition = q.order_by(TuitionInfo.id.desc()).first()
    if not tuition:
        return jsonify({"message": "No tuition info found"}), 404

    return jsonify({
        "id": tuition.id,
        "student_id": tuition.student_id,
        "term": tuition.term,
        "total_amount": tuition.total_amount,
        "amount_paid": tuition.amount_paid,
        "balance": float(tuition.total_amount) - float(tuition.amount_paid),
        "balance_due": float(tuition.total_amount) - float(tuition.amount_paid),
        "payment_plan": tuition.payment_plan,
        "status": tuition.status,
        "payments": [
            {
                "id": p.id,
                "amount": p.amount,
                "method": p.method,
                "reference": p.reference,
                "timestamp": p.timestamp.isoformat() if p.timestamp else None,
                "note": p.note,
            }
            for p in tuition.payments
        ],
    }), 200


@finance_bp.route("/api/s/<slug>/tuition", methods=["POST"])
@finance_bp.route("/api/tuition", methods=["POST"])
@school_context_required
@roles_required(ROLE_ADMIN)
@bridge.limiter.limit("60 per hour")
def create_or_update_tuition(slug=None):
    Student = bridge.Student
    TuitionInfo = bridge.TuitionInfo
    db = bridge.db

    data = request.get_json(silent=True) or {}
    required = ["student_id", "term", "total_amount"]
    missing = [k for k in required if k not in data]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400

    student_id = int(data["student_id"])
    term = str(data["term"]).strip()
    total_amount = float(data["total_amount"])

    student = Student.query.filter_by(
        id=student_id,
        school_id=current_school_id()
    ).first()

    if not student:
        return jsonify({"error": "Student not found"}), 404

    tuition = TuitionInfo.query.filter_by(
        school_id=current_school_id(),
        student_id=student_id,
        term=term
    ).first()

    if not tuition:
        tuition = TuitionInfo(
            school_id=current_school_id(),
            student_id=student_id,
            term=term,
            total_amount=total_amount
        )
        db.session.add(tuition)
    else:
        tuition.total_amount = total_amount

    if "amount_paid" in data:
        tuition.amount_paid = float(data["amount_paid"])
    tuition.payment_plan = data.get("payment_plan")
    tuition.status = data.get("status")

    db.session.commit()

    log_audit(
        module="finance",
        action="save_tuition",
        entity_type="tuition",
        entity_id=tuition.id,
        entity_label=f"{student.name} - {term}",
        details={
            "student_id": student.id,
            "student_name": student.name,
            "term": term,
            "total_amount": tuition.total_amount,
            "amount_paid": tuition.amount_paid,
            "status": tuition.status,
            "payment_plan": tuition.payment_plan,
        },
    )

    return jsonify({"id": tuition.id}), 200


@finance_bp.route("/api/s/<slug>/tuition/<int:tuition_id>/payment", methods=["POST"])
@finance_bp.route("/api/tuition/<int:tuition_id>/payment", methods=["POST"])
@school_context_required
@roles_required(ROLE_ADMIN)
@bridge.limiter.limit("60 per hour")
def add_payment(tuition_id, slug=None):
    TuitionInfo = bridge.TuitionInfo
    PaymentHistory = bridge.PaymentHistory
    db = bridge.db

    tuition = TuitionInfo.query.filter_by(
        id=tuition_id,
        school_id=current_school_id()
    ).first()

    if not tuition:
        return jsonify({"error": "Tuition record not found"}), 404

    data = request.get_json(silent=True) or {}
    if "amount" not in data:
        return jsonify({"error": "amount is required"}), 400

    p = PaymentHistory(
        school_id=current_school_id(),
        tuition_id=tuition.id,
        amount=float(data["amount"]),
        method=data.get("method"),
        reference=data.get("reference"),
        note=data.get("note"),
    )
    db.session.add(p)

    tuition.amount_paid = float(tuition.amount_paid or 0) + float(p.amount)
    db.session.commit()

    log_audit(
        module="finance",
        action="add_payment",
        entity_type="payment",
        entity_id=p.id,
        entity_label=f"Payment for student {tuition.student_id}",
        details={
            "tuition_id": tuition.id,
            "student_id": tuition.student_id,
            "amount": p.amount,
            "method": p.method,
            "reference": p.reference,
        },
    )

    return jsonify({"payment_id": p.id, "amount_paid": tuition.amount_paid}), 201


@finance_bp.route("/api/s/<slug>/payments/<int:student_id>", methods=["GET"])
@finance_bp.route("/api/payments/<int:student_id>", methods=["GET"])
@school_context_required
@roles_required(ROLE_ADMIN)
def api_payments_by_student(student_id, slug=None):
    Student = bridge.Student
    PaymentHistory = bridge.PaymentHistory
    TuitionInfo = bridge.TuitionInfo

    student = Student.query.filter_by(
        id=student_id,
        school_id=current_school_id()
    ).first()

    if not student:
        return jsonify({"error": "Student not found"}), 404

    term = request.args.get("term")
    q = (
        PaymentHistory.query
        .join(TuitionInfo, TuitionInfo.id == PaymentHistory.tuition_id)
        .filter(
            TuitionInfo.school_id == current_school_id(),
            TuitionInfo.student_id == student_id
        )
    )

    if term:
        q = q.filter(TuitionInfo.term == term)

    payments = q.order_by(PaymentHistory.timestamp.desc()).all()

    return jsonify([
        {
            "id": p.id,
            "student_id": student_id,
            "tuition_id": p.tuition_id,
            "amount": float(p.amount or 0),
            "method": p.method,
            "reference": p.reference,
            "note": p.note,
            "timestamp": p.timestamp.isoformat() if p.timestamp else None,
        }
        for p in payments
    ]), 200


@finance_bp.route("/api/s/<slug>/finance/summary", methods=["GET"])
@finance_bp.route("/api/finance/summary", methods=["GET"])
@school_context_required
@roles_required(ROLE_ADMIN)
@bridge.limiter.limit("30 per hour")
def finance_summary(slug=None):
    TuitionInfo = bridge.TuitionInfo

    term = request.args.get("term")

    q = TuitionInfo.query.filter_by(school_id=current_school_id())
    if term:
        q = q.filter(TuitionInfo.term == term)

    rows = q.all()

    total_billed = sum(_money(r.total_amount) for r in rows)
    total_paid = sum(_money(r.amount_paid) for r in rows)
    total_balance = total_billed - total_paid

    return jsonify({
        "term": term,
        "total_students": len(rows),
        "total_billed": total_billed,
        "total_paid": total_paid,
        "total_balance": total_balance,
    }), 200


@finance_bp.route("/api/s/<slug>/finance/dashboard", methods=["GET"])
@finance_bp.route("/api/finance/dashboard", methods=["GET"])
@school_context_required
@roles_required(ROLE_ADMIN)
@bridge.limiter.limit("30 per hour")
def finance_dashboard(slug=None):
    TuitionInfo = bridge.TuitionInfo
    Student = bridge.Student

    term = request.args.get("term")

    q = TuitionInfo.query.filter_by(school_id=current_school_id())

    if term:
        q = q.filter(TuitionInfo.term == term)

    rows = q.all()

    student_ids = [r.student_id for r in rows if r.student_id]
    students = Student.query.filter(
        Student.school_id == current_school_id(),
        Student.id.in_(student_ids) if student_ids else False
    ).all() if student_ids else []

    student_map = {s.id: s for s in students}

    items = []
    for r in rows:
        student = student_map.get(r.student_id)
        balance = _money(r.total_amount) - _money(r.amount_paid)

        items.append({
            "tuition_id": r.id,
            "student_id": r.student_id,
            "student_name": student.name if student else None,
            "grade": student.grade if student else None,
            "term": r.term,
            "total_amount": _money(r.total_amount),
            "amount_paid": _money(r.amount_paid),
            "balance": balance,
            "status": r.status,
            "payment_plan": r.payment_plan,
        })

    total_billed = sum(i["total_amount"] for i in items)
    total_paid = sum(i["amount_paid"] for i in items)
    total_balance = sum(i["balance"] for i in items)

    return jsonify({
        "summary": {
            "count": len(items),
            "total_billed": total_billed,
            "total_paid": total_paid,
            "total_balance": total_balance,
        },
        "items": items,
    }), 200


@finance_bp.route("/api/s/<slug>/finance/statement", methods=["GET"])
@finance_bp.route("/api/finance/statement", methods=["GET"])
@school_context_required
@roles_required(ROLE_ADMIN)
@bridge.limiter.limit("30 per hour")
def finance_statement_export(slug=None):
    TuitionInfo = bridge.TuitionInfo
    Student = bridge.Student

    term = request.args.get("term")

    q = TuitionInfo.query.filter_by(school_id=current_school_id())

    if term:
        q = q.filter(TuitionInfo.term == term)

    rows = q.all()

    student_ids = [r.student_id for r in rows if r.student_id]
    students = Student.query.filter(
        Student.school_id == current_school_id(),
        Student.id.in_(student_ids) if student_ids else False
    ).all() if student_ids else []

    student_map = {s.id: s for s in students}

    items = []
    for r in rows:
        student = student_map.get(r.student_id)

        items.append({
            "student_id": r.student_id,
            "student_name": student.name if student else "",
            "grade": student.grade if student else "",
            "term": r.term or "",
            "total_amount": _money(r.total_amount),
            "amount_paid": _money(r.amount_paid),
            "balance": _money(r.total_amount) - _money(r.amount_paid),
            "status": r.status or "",
            "payment_plan": r.payment_plan or "",
        })

    log_audit(
        module="finance",
        action="export_statement",
        entity_type="finance_statement",
        entity_id=term or "all",
        entity_label="Finance Statement Export",
        details={"term": term, "exported_count": len(items)},
    )

    return jsonify({
        "term": term,
        "exported_count": len(items),
        "items": items,
    }), 200

@finance_bp.route("/api/s/<slug>/finance/import", methods=["POST"])
@finance_bp.route("/api/finance/import", methods=["POST"])
@school_context_required
@roles_required(ROLE_ADMIN)
@bridge.limiter.limit("10 per hour")
def import_finance(slug=None):
    Student = bridge.Student
    TuitionInfo = bridge.TuitionInfo
    PaymentHistory = bridge.PaymentHistory
    db = bridge.db
    sid = current_school_id()

    kind = (request.form.get("kind") or "").strip().lower()
    file = request.files.get("file")

    if kind not in ("tuition", "payment"):
        return jsonify({"error": "kind must be tuition or payment"}), 400

    if not file or not file.filename:
        return jsonify({"error": "Import file is required"}), 400

    try:
        file_type, headers, import_rows = _load_import_rows(file)
    except UnicodeDecodeError:
        return jsonify({"error": "Could not read CSV. Please save it as UTF-8 CSV."}), 400
    except Exception as e:
        return jsonify({"error": str(e) or "Could not read import file"}), 400

    if kind == "tuition":
        required = ["student_id", "term", "total_amount"]
    else:
        required = ["student_id", "term", "amount"]

    normalized_headers = {h: h for h in headers if h}
    missing = [h for h in required if h not in normalized_headers]

    if missing:
        return jsonify({
            "error": f"Missing required column(s): {', '.join(missing)}",
            "kind": kind,
            "required_columns": required,
            "tuition_columns": [
                "student_id",
                "term",
                "total_amount",
                "amount_paid",
                "payment_plan",
                "status",
            ],
            "payment_columns": [
                "student_id",
                "term",
                "amount",
                "method",
                "reference",
                "note",
            ],
        }), 400

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for index, row in enumerate(import_rows, start=2):
        row = {str(k).strip().lower(): v for k, v in row.items() if k is not None}

        student_id_raw = _clean_cell(row, "student_id")
        term = _clean_cell(row, "term")

        if not student_id_raw or not term:
            skipped += 1
            errors.append({
                "row": index,
                "error": "student_id and term are required",
            })
            continue

        try:
            student_id = int(float(student_id_raw))
        except Exception:
            skipped += 1
            errors.append({
                "row": index,
                "student_id": student_id_raw,
                "error": "student_id must be a number",
            })
            continue

        student = Student.query.filter_by(
            id=student_id,
            school_id=sid
        ).first()

        if not student:
            skipped += 1
            errors.append({
                "row": index,
                "student_id": student_id,
                "error": "Student not found",
            })
            continue

        tuition = TuitionInfo.query.filter_by(
            school_id=sid,
            student_id=student_id,
            term=term
        ).first()

        if kind == "tuition":
            total_amount = _to_float(_clean_cell(row, "total_amount"))
            amount_paid = _to_float(_clean_cell(row, "amount_paid"))
            payment_plan = _clean_cell(row, "payment_plan") or None
            status = _clean_cell(row, "status") or None

            if not tuition:
                tuition = TuitionInfo(
                    school_id=sid,
                    student_id=student_id,
                    term=term,
                    total_amount=total_amount,
                    amount_paid=amount_paid,
                    payment_plan=payment_plan,
                    status=status,
                )
                db.session.add(tuition)
                created += 1
            else:
                tuition.total_amount = total_amount
                tuition.amount_paid = amount_paid
                tuition.payment_plan = payment_plan
                tuition.status = status
                updated += 1

        else:
            amount = _to_float(_clean_cell(row, "amount"))
            method = _clean_cell(row, "method") or None
            reference = _clean_cell(row, "reference") or None
            note = _clean_cell(row, "note") or None

            if amount <= 0:
                skipped += 1
                errors.append({
                    "row": index,
                    "student_id": student_id,
                    "term": term,
                    "error": "amount must be greater than 0",
                })
                continue

            if not tuition:
                tuition = TuitionInfo(
                    school_id=sid,
                    student_id=student_id,
                    term=term,
                    total_amount=0,
                    amount_paid=0,
                    payment_plan=None,
                    status=None,
                )
                db.session.add(tuition)
                db.session.flush()
                created += 1

            payment = PaymentHistory(
                school_id=sid,
                tuition_id=tuition.id,
                amount=amount,
                method=method,
                reference=reference,
                note=note,
            )
            db.session.add(payment)

            tuition.amount_paid = float(tuition.amount_paid or 0) + amount
            updated += 1

    db.session.commit()

    log_audit(
        module="finance",
        action="bulk_import",
        entity_type="finance_import",
        entity_id=None,
        entity_label="Bulk Finance Import",
        details={
            "kind": kind,
            "file_type": file_type,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors_count": len(errors),
        },
    )

    return jsonify({
        "message": "Finance import completed",
        "kind": kind,
        "file_type": file_type,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:100],
    }), 200

@finance_bp.route("/api/s/<slug>/payments/<int:payment_id>/receipt.pdf", methods=["GET"])
@finance_bp.route("/api/payments/<int:payment_id>/receipt/pdf", methods=["GET"])
@finance_bp.route("/api/payments/<int:payment_id>/receipt.pdf", methods=["GET"])
@school_context_required
@roles_required(ROLE_ADMIN)
@bridge.limiter.limit("30 per hour")
def payment_receipt_pdf(payment_id, slug=None):
    PaymentHistory = bridge.PaymentHistory

    payment = PaymentHistory.query.filter_by(
        id=payment_id,
        school_id=current_school_id()
    ).first()

    if not payment:
        return jsonify({"error": "Payment not found"}), 404

    log_audit(
        module="finance",
        action="open_receipt_pdf",
        entity_type="payment",
        entity_id=payment.id,
        entity_label=f"Receipt {payment.id}",
        details={"amount": payment.amount, "method": payment.method},
    )

    content = (
        f"Receipt\n\n"
        f"Payment ID: {payment.id}\n"
        f"Amount: {_money(payment.amount):.2f}\n"
        f"Method: {payment.method or ''}\n"
        f"Reference: {payment.reference or ''}\n"
        f"Note: {payment.note or ''}\n"
    ).encode("utf-8")

    return send_file(
        io.BytesIO(content),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"receipt_{payment.id}.pdf",
    )