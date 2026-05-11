from flask import Blueprint, request, jsonify, url_for, send_from_directory
from werkzeug.utils import secure_filename
import os
import csv
import io
from datetime import datetime
from openpyxl import load_workbook
from app.audit import log_audit
import app.bridge as bridge
from app.decorators import (
    ROLE_ADMIN,
    ROLE_TEACHER,
    ROLE_PARENT,
    ROLE_STUDENT,
    current_school_id,
    is_admin,
    roles_required,
    school_context_required,
)

students_bp = Blueprint("students_bp", __name__)

def _normalize_header(value):
    return str(value or "").strip().lower()


def _rows_from_csv(file):
    raw = file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))

    if not reader.fieldnames:
        raise ValueError("CSV has no header row")

    rows = []
    for row in reader:
        rows.append({str(k).strip().lower(): v for k, v in row.items() if k is not None})

    return [_normalize_header(h) for h in reader.fieldnames if h], rows


def _rows_from_xlsx(file):
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        raise ValueError("Excel file is empty")

    headers = [_normalize_header(h) for h in rows_iter[0]]
    if not any(headers):
        raise ValueError("Excel file has no header row")

    rows = []
    for values in rows_iter[1:]:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[idx] if idx < len(values) else ""
        rows.append(row)

    return headers, rows

def _parse_import_date(value):
    raw = str(value or "").strip()
    if not raw:
        return None

    # Try your existing app parser first
    try:
        return bridge.parse_date(raw)
    except Exception:
        pass

    # Common spreadsheet/export formats
    formats = [
        "%Y-%m-%d",   # 2015-05-10
        "%m/%d/%Y",   # 05/10/2015
        "%m/%d/%y",   # 05/10/15
        "%d/%m/%Y",   # 10/05/2015
        "%d/%m/%y",   # 10/05/15
        "%Y/%m/%d",   # 2015/05/10
        "%m-%d-%Y",   # 05-10-2015
        "%d-%m-%Y",   # 10-05-2015
    ]

    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt).date()
        except Exception:
            continue

    raise ValueError("Invalid date")

def _student_to_dict(student):
    return {
        "id": student.id,
        "name": student.name,
        "dob": student.dob.isoformat() if student.dob else None,
        "gender": student.gender,
        "national_id": student.national_id,
        "grade": student.grade,
        "email": student.email,
        "guardian_name": student.guardian_name,
        "guardian_contact": student.guardian_contact,
        "home_address": student.home_address,
        "emergency_contact": student.emergency_contact,
        "photo_url": (
            url_for("students_bp.serve_photo", filename=student.photo_filename, _external=True)
            if student.photo_filename
            else None
        ),
    }


def _clean_cell(row, key, default=""):
    value = row.get(key, default)
    if value is None:
        return default
    return str(value).strip()


def _parse_grade(value):
    try:
        return int(str(value or "1").strip())
    except Exception:
        return 1


@students_bp.route("/api/s/<slug>/students", methods=["GET", "POST"])
@students_bp.route("/api/students", methods=["GET", "POST"])
@school_context_required
@roles_required(ROLE_ADMIN, ROLE_TEACHER)
def students(slug=None):
    Student = bridge.Student
    db = bridge.db

    if request.method == "GET":
        rows = Student.query.filter_by(
            school_id=current_school_id()
        ).order_by(Student.id).all()
        return jsonify([_student_to_dict(s) for s in rows]), 200

    if not is_admin():
        return jsonify({"error": "Forbidden"}), 403

    data = request.get_json(silent=True) or {}
    if not data.get("name"):
        return jsonify({"error": "Name required"}), 400

    s = Student(
        school_id=current_school_id(),
        name=data["name"].strip(),
        dob=bridge.parse_date(data["dob"]) if data.get("dob") else None,
        gender=data.get("gender"),
        national_id=data.get("national_id"),
        grade=int(data.get("grade", 1)),
        email=data.get("email"),
        guardian_name=data.get("guardian_name"),
        guardian_contact=data.get("guardian_contact"),
        home_address=data.get("home_address"),
        emergency_contact=data.get("emergency_contact"),
    )
    db.session.add(s)
    db.session.commit()

    log_audit(
        module="students",
        action="create",
        entity_type="student",
        entity_id=s.id,
        entity_label=s.name,
        details={"grade": s.grade, "guardian_name": s.guardian_name},
    )

    return jsonify({"id": s.id}), 201


@students_bp.route("/api/s/<slug>/students/import", methods=["POST"])
@students_bp.route("/api/students/import", methods=["POST"])
@school_context_required
@roles_required(ROLE_ADMIN)
@bridge.limiter.limit("10 per hour")
def import_students_csv(slug=None):
    Student = bridge.Student
    db = bridge.db
    sid = current_school_id()

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "CSV file is required"}), 400

    filename = file.filename.lower()
    if not filename.endswith(".csv"):
        return jsonify({"error": "Only .csv files are supported for now"}), 400

    try:
        raw = file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        return jsonify({"error": "Could not read CSV. Please save it as UTF-8 CSV."}), 400

    reader = csv.DictReader(io.StringIO(raw))

    if not reader.fieldnames:
        return jsonify({"error": "CSV has no header row"}), 400

    normalized_headers = {h.strip().lower(): h for h in reader.fieldnames if h}
    required = ["name", "grade"]
    missing = [h for h in required if h not in normalized_headers]

    if missing:
        return jsonify({
            "error": f"Missing required column(s): {', '.join(missing)}",
            "required_columns": required,
            "optional_columns": [
                "dob",
                "gender",
                "email",
                "national_id",
                "guardian_name",
                "guardian_contact",
                "home_address",
                "emergency_contact",
            ],
        }), 400

    created = 0
    skipped = 0
    errors = []
    created_students = []

    for index, row in enumerate(import_rows, start=2):
        row = {str(k).strip().lower(): v for k, v in row.items() if k is not None}

        name = _clean_cell(row, "name")
        grade = _parse_grade(_clean_cell(row, "grade", "1"))

        if not name:
            skipped += 1
            errors.append({
                "row": index,
                "error": "Missing student name",
            })
            continue

        existing = Student.query.filter_by(
            school_id=sid,
            name=name,
            grade=grade,
        ).first()

        if existing:
            skipped += 1
            errors.append({
                "row": index,
                "name": name,
                "grade": grade,
                "error": "Duplicate student skipped",
            })
            continue

        dob_raw = _clean_cell(row, "dob")
        dob = None

        if dob_raw:
            try:
                dob = _parse_import_date(dob_raw)
            except Exception:
                skipped += 1
                errors.append({
                    "row": index,
                    "name": name,
                    "error": "Invalid dob. Use YYYY-MM-DD, MM/DD/YYYY, or DD/MM/YYYY.",
                })
                continue

        student = Student(
            school_id=sid,
            name=name,
            dob=dob,
            gender=_clean_cell(row, "gender") or None,
            national_id=_clean_cell(row, "national_id") or None,
            grade=grade,
            email=_clean_cell(row, "email") or None,
            guardian_name=_clean_cell(row, "guardian_name") or None,
            guardian_contact=_clean_cell(row, "guardian_contact") or None,
            home_address=_clean_cell(row, "home_address") or None,
            emergency_contact=_clean_cell(row, "emergency_contact") or None,
        )

        db.session.add(student)
        created_students.append(student)
        created += 1

    db.session.commit()

    log_audit(
        module="students",
        action="bulk_import",
        entity_type="student_import",
        entity_id=None,
        entity_label="Bulk Student Import",
        details={
            "created": created,
            "file_type": "xlsx" if filename.endswith(".xlsx") else "csv",
            "skipped": skipped,
            "errors_count": len(errors),
            "sample_created": [
                {
                    "id": s.id,
                    "name": s.name,
                    "grade": s.grade,
                }
                for s in created_students[:10]
            ],
        },
    )

    return jsonify({
        "message": "Import completed",
        "created": created,
        "skipped": skipped,
        "errors": errors[:100],
        "created_students": [_student_to_dict(s) for s in created_students[:50]],
    }), 200


@students_bp.route("/api/s/<slug>/students/<int:student_id>", methods=["GET", "PUT", "DELETE"])
@students_bp.route("/api/students/<int:student_id>", methods=["GET", "PUT", "DELETE"])
@school_context_required
@roles_required(ROLE_ADMIN, ROLE_TEACHER, ROLE_PARENT, ROLE_STUDENT)
def modify_student(student_id, slug=None):
    Student = bridge.Student
    db = bridge.db

    s = Student.query.filter_by(
        id=student_id,
        school_id=current_school_id()
    ).first()

    if not s:
        return jsonify({"error": "Student not found"}), 404

    if request.method == "GET":
        return jsonify(_student_to_dict(s)), 200

    if request.method in ["PUT", "DELETE"] and not is_admin():
        return jsonify({"error": "Forbidden"}), 403

    if request.method == "PUT":
        data = request.get_json(silent=True) or {}

        for fld in [
            "name",
            "gender",
            "national_id",
            "email",
            "guardian_name",
            "guardian_contact",
            "home_address",
            "emergency_contact",
        ]:
            if fld in data:
                setattr(s, fld, data[fld])

        if "dob" in data:
            s.dob = bridge.parse_date(data["dob"]) if data["dob"] else None
        if "grade" in data:
            s.grade = int(data["grade"])

        db.session.commit()

        log_audit(
            module="students",
            action="update",
            entity_type="student",
            entity_id=s.id,
            entity_label=s.name,
            details={"updated_fields": list(data.keys())},
        )

        return jsonify({"message": "Student updated"}), 200

    student_name = s.name
    student_grade = s.grade

    db.session.delete(s)
    db.session.commit()

    log_audit(
        module="students",
        action="delete",
        entity_type="student",
        entity_id=student_id,
        entity_label=student_name,
        details={"grade": student_grade},
    )

    return jsonify({"message": "Student deleted"}), 200


@students_bp.route("/api/s/<slug>/students/<int:student_id>/photo", methods=["POST"])
@students_bp.route("/api/students/<int:student_id>/photo", methods=["POST"])
@school_context_required
@roles_required(ROLE_ADMIN)
@bridge.limiter.limit("20 per hour")
def upload_student_photo(student_id, slug=None):
    Student = bridge.Student
    db = bridge.db

    s = Student.query.filter_by(
        id=student_id,
        school_id=current_school_id()
    ).first()

    if not s:
        return jsonify({"error": "Student not found"}), 404

    file = request.files.get("photo")
    if not file or file.filename == "":
        return jsonify({"error": "No file"}), 400

    if not bridge.allowed_file(file.filename, bridge.ALLOWED_PHOTO):
        return jsonify({"error": "Invalid type"}), 400

    upload_photos_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        "uploads",
        "photos",
    )
    os.makedirs(upload_photos_dir, exist_ok=True)

    fname = secure_filename(f"{student_id}_{int(datetime.utcnow().timestamp())}_{file.filename}")
    file.save(os.path.join(upload_photos_dir, fname))

    s.photo_filename = fname
    db.session.commit()

    log_audit(
        module="students",
        action="upload_photo",
        entity_type="student",
        entity_id=s.id,
        entity_label=s.name,
        details={"photo_filename": fname},
    )

    return jsonify({
        "message": "Photo uploaded",
        "photo_url": url_for("students_bp.serve_photo", filename=fname, _external=True)
    }), 201


@students_bp.route("/photos/<filename>")
def serve_photo(filename):
    upload_photos_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        "uploads",
        "photos",
    )
    return send_from_directory(upload_photos_dir, filename)